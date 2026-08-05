"""Pruebas de bitacora.py - resolver de Escritorio, excel de operaciones
(balances diario/semanal/mensual + operativas) y resumenes diarios en texto.

El excel y los resumenes se prueban pasando tmp_path directo como "raiz"
(mismo patron que memoria.py/test_memoria.py) - nunca tocan el Escritorio
real. El resolver de Escritorio mockea winreg directamente y nunca toca el
registro real.
"""
import winreg
from datetime import date, timedelta
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

import bitacora


def _operacion(fecha, resultado_usd, mercado="BTC/USDT", tipo="Long", operativa="Breakout", hora="10:00"):
    return {
        "fecha": fecha,
        "hora": hora,
        "mercado": mercado,
        "operativa": operativa,
        "tipo": tipo,
        "entrada": 100.0,
        "salida": 110.0,
        "tamano": 1.0,
        "resultado_usd": resultado_usd,
        "resultado_r": 1.0,
        "notas": "",
    }


def _fecha_y_siguiente_misma_semana_distinto_mes():
    """Encuentra un par de fechas consecutivas que compartan semana ISO pero
    esten en meses calendario distintos - para probar que Balance Semanal y
    Balance Mensual agrupan de forma independiente, sin hardcodear un par
    de fechas que podria dejar de servir si cambia el codigo."""
    fecha = date(2026, 1, 1)
    while True:
        siguiente = fecha + timedelta(days=1)
        if (
            fecha.isocalendar()[:2] == siguiente.isocalendar()[:2]
            and fecha.month != siguiente.month
        ):
            return fecha, siguiente
        fecha = siguiente


class _FakeClave:
    def __enter__(self):
        return self

    def __exit__(self, *args):
        return False


# --- Normalizacion -----------------------------------------------------

def test_normalizar_operacion_fecha_por_defecto_hoy():
    op = bitacora._normalizar_operacion({"resultado_usd": 10, "mercado": "BTC", "tipo": "Long"})
    assert op["fecha"] == date.today()


def test_normalizar_operacion_falta_resultado_usd():
    with pytest.raises(ValueError):
        bitacora._normalizar_operacion({"mercado": "BTC"})


def test_normalizar_operacion_fecha_invalida():
    with pytest.raises(ValueError):
        bitacora._normalizar_operacion({"resultado_usd": 5, "fecha": "no-es-fecha"})


def test_normalizar_operacion_campo_opcional_invalido_no_bloquea():
    op = bitacora._normalizar_operacion({"resultado_usd": 5, "tamano": "abc"})
    assert op["tamano"] is None


def test_normalizar_operacion_resultado_cero_es_valido():
    op = bitacora._normalizar_operacion({"resultado_usd": 0, "mercado": "BTC", "tipo": "Long"})
    assert op["resultado_usd"] == 0.0


# --- Excel: creacion y append -------------------------------------------

def test_guardar_operacion_crea_las_5_hojas(tmp_path):
    bitacora._guardar_operacion_excel(tmp_path, _operacion(date(2026, 8, 1), 50))
    wb = load_workbook(bitacora._ruta_excel(tmp_path))
    assert set(wb.sheetnames) == {
        bitacora._HOJA_OPERACIONES,
        bitacora._HOJA_BALANCE_DIARIO,
        bitacora._HOJA_BALANCE_SEMANAL,
        bitacora._HOJA_BALANCE_MENSUAL,
        bitacora._HOJA_OPERATIVAS,
    }


def test_guardar_operacion_agrega_fila_sin_tocar_anteriores(tmp_path):
    bitacora._guardar_operacion_excel(tmp_path, _operacion(date(2026, 8, 1), 50))
    bitacora._guardar_operacion_excel(tmp_path, _operacion(date(2026, 8, 2), -20))
    wb = load_workbook(bitacora._ruta_excel(tmp_path))
    assert wb[bitacora._HOJA_OPERACIONES].max_row == 3  # encabezado + 2 filas


# --- Balances -------------------------------------------------------------

def test_balance_diario_agrupa_mismo_dia(tmp_path):
    bitacora._guardar_operacion_excel(tmp_path, _operacion(date(2026, 8, 1), 50))
    bitacora._guardar_operacion_excel(tmp_path, _operacion(date(2026, 8, 1), -20))
    bitacora._guardar_operacion_excel(tmp_path, _operacion(date(2026, 8, 2), 10))

    wb = load_workbook(bitacora._ruta_excel(tmp_path))
    filas = list(wb[bitacora._HOJA_BALANCE_DIARIO].iter_rows(min_row=2, values_only=True))
    assert len(filas) == 2
    fila_dia_1 = next(f for f in filas if f[0] == "2026-08-01")
    assert fila_dia_1[1] == 2
    assert fila_dia_1[2] == 30


def test_balance_semanal_agrupa_misma_semana_iso_distinto_mes(tmp_path):
    fecha_a, fecha_b = _fecha_y_siguiente_misma_semana_distinto_mes()
    bitacora._guardar_operacion_excel(tmp_path, _operacion(fecha_a, 50))
    bitacora._guardar_operacion_excel(tmp_path, _operacion(fecha_b, 30))

    wb = load_workbook(bitacora._ruta_excel(tmp_path))
    filas = list(wb[bitacora._HOJA_BALANCE_SEMANAL].iter_rows(min_row=2, values_only=True))
    assert len(filas) == 1
    assert filas[0][1] == 2
    assert filas[0][2] == 80


def test_balance_mensual_separa_por_mes_aunque_misma_semana_iso(tmp_path):
    fecha_a, fecha_b = _fecha_y_siguiente_misma_semana_distinto_mes()
    bitacora._guardar_operacion_excel(tmp_path, _operacion(fecha_a, 50))
    bitacora._guardar_operacion_excel(tmp_path, _operacion(fecha_b, 30))

    wb = load_workbook(bitacora._ruta_excel(tmp_path))
    filas = list(wb[bitacora._HOJA_BALANCE_MENSUAL].iter_rows(min_row=2, values_only=True))
    assert len(filas) == 2


def test_operativas_agrupa_y_calcula_winrate(tmp_path):
    bitacora._guardar_operacion_excel(tmp_path, _operacion(date(2026, 8, 1), 50, operativa="Breakout"))
    bitacora._guardar_operacion_excel(tmp_path, _operacion(date(2026, 8, 2), 30, operativa="Breakout"))
    bitacora._guardar_operacion_excel(tmp_path, _operacion(date(2026, 8, 3), -20, operativa="Breakout"))
    bitacora._guardar_operacion_excel(tmp_path, _operacion(date(2026, 8, 4), 15, operativa="Pullback"))

    wb = load_workbook(bitacora._ruta_excel(tmp_path))
    filas = {f[0]: f for f in wb[bitacora._HOJA_OPERATIVAS].iter_rows(min_row=2, values_only=True)}
    assert filas["Breakout"][1] == 3
    assert filas["Breakout"][2] == 60
    assert filas["Breakout"][3] == pytest.approx(66.7, abs=0.1)
    assert filas["Pullback"][1] == 1


# --- Racha de perdidas -----------------------------------------------------

def test_racha_perdidas_sin_operaciones_hoy_es_cero():
    assert bitacora._racha_perdidas_consecutivas([], date(2026, 8, 4)) == 0


def test_racha_perdidas_una_sola_perdida_es_uno():
    operaciones = [_operacion(date(2026, 8, 4), -10, hora="09:00")]
    assert bitacora._racha_perdidas_consecutivas(operaciones, date(2026, 8, 4)) == 1


def test_racha_perdidas_tres_seguidas_es_tres():
    operaciones = [
        _operacion(date(2026, 8, 4), -10, hora="09:00"),
        _operacion(date(2026, 8, 4), -5, hora="10:00"),
        _operacion(date(2026, 8, 4), -20, hora="11:00"),
    ]
    assert bitacora._racha_perdidas_consecutivas(operaciones, date(2026, 8, 4)) == 3


def test_racha_perdidas_gana_resetea_y_siguiente_perdida_arranca_de_uno():
    operaciones = [
        _operacion(date(2026, 8, 4), -10, hora="09:00"),
        _operacion(date(2026, 8, 4), -5, hora="09:30"),
        _operacion(date(2026, 8, 4), 15, hora="10:00"),
    ]
    assert bitacora._racha_perdidas_consecutivas(operaciones, date(2026, 8, 4)) == 0

    operaciones.append(_operacion(date(2026, 8, 4), -8, hora="10:30"))
    assert bitacora._racha_perdidas_consecutivas(operaciones, date(2026, 8, 4)) == 1


def test_racha_perdidas_empate_corta_racha_igual_que_ganar():
    operaciones = [
        _operacion(date(2026, 8, 4), -10, hora="09:00"),
        _operacion(date(2026, 8, 4), -5, hora="09:30"),
        _operacion(date(2026, 8, 4), 0, hora="10:00"),
    ]
    assert bitacora._racha_perdidas_consecutivas(operaciones, date(2026, 8, 4)) == 0


def test_racha_perdidas_no_cuenta_dia_anterior():
    operaciones = [
        _operacion(date(2026, 8, 3), -10, hora="09:00"),
        _operacion(date(2026, 8, 3), -5, hora="10:00"),
        _operacion(date(2026, 8, 4), -8, hora="09:00"),
    ]
    assert bitacora._racha_perdidas_consecutivas(operaciones, date(2026, 8, 4)) == 1


def test_racha_perdidas_ordena_por_hora_no_por_orden_de_lista():
    # Pasadas en orden no cronologico a proposito - si la funcion no
    # ordenara por hora, tomaria la ultima de la lista (perdida) en vez de
    # la ultima cronologica (ganancia) y daria 1 en vez de 0.
    operaciones = [
        _operacion(date(2026, 8, 4), -10, hora="10:00"),
        _operacion(date(2026, 8, 4), 20, hora="11:00"),
        _operacion(date(2026, 8, 4), -5, hora="09:00"),
    ]
    assert bitacora._racha_perdidas_consecutivas(operaciones, date(2026, 8, 4)) == 0


def test_racha_perdidas_del_dia_lee_desde_excel_guardado(tmp_path):
    fecha = date(2026, 8, 4)
    bitacora._guardar_operacion_excel(tmp_path, _operacion(fecha, -10, hora="09:00"))
    bitacora._guardar_operacion_excel(tmp_path, _operacion(fecha, -5, hora="10:00"))
    assert bitacora._racha_perdidas_del_dia(tmp_path, fecha) == 2


def test_registrar_operacion_perdida_aislada_no_agrega_alerta(tmp_path, monkeypatch):
    monkeypatch.setattr(bitacora, "_raiz_bitacora", lambda: tmp_path)
    resultado = bitacora.registrar_operacion({"mercado": "BTC", "tipo": "Long", "resultado_usd": -10})
    assert "perdidas seguidas" not in resultado.lower()


def test_registrar_operacion_racha_dos_perdidas_agrega_alerta(tmp_path, monkeypatch):
    monkeypatch.setattr(bitacora, "_raiz_bitacora", lambda: tmp_path)
    bitacora.registrar_operacion({"mercado": "BTC", "tipo": "Long", "resultado_usd": -10})
    resultado = bitacora.registrar_operacion({"mercado": "BTC", "tipo": "Long", "resultado_usd": -5})
    assert "perdidas seguidas" in resultado.lower()


def test_registrar_operacion_racha_tres_menciona_el_numero(tmp_path, monkeypatch):
    monkeypatch.setattr(bitacora, "_raiz_bitacora", lambda: tmp_path)
    bitacora.registrar_operacion({"mercado": "BTC", "tipo": "Long", "resultado_usd": -10})
    bitacora.registrar_operacion({"mercado": "BTC", "tipo": "Long", "resultado_usd": -5})
    resultado = bitacora.registrar_operacion({"mercado": "BTC", "tipo": "Long", "resultado_usd": -8})
    assert "van 3 perdidas seguidas" in resultado.lower()


def test_registrar_operacion_ganancia_no_agrega_alerta(tmp_path, monkeypatch):
    monkeypatch.setattr(bitacora, "_raiz_bitacora", lambda: tmp_path)
    bitacora.registrar_operacion({"mercado": "BTC", "tipo": "Long", "resultado_usd": -10})
    bitacora.registrar_operacion({"mercado": "BTC", "tipo": "Long", "resultado_usd": -5})
    resultado = bitacora.registrar_operacion({"mercado": "BTC", "tipo": "Long", "resultado_usd": 20})
    assert "perdidas seguidas" not in resultado.lower()


# --- Regresiones de la reescritura de hojas de resumen ---------------------

def test_recompute_no_modifica_operaciones(tmp_path):
    bitacora._guardar_operacion_excel(tmp_path, _operacion(date(2026, 8, 1), 50))
    bitacora._guardar_operacion_excel(tmp_path, _operacion(date(2026, 8, 2), -20))
    filas_antes = list(
        load_workbook(bitacora._ruta_excel(tmp_path))[bitacora._HOJA_OPERACIONES].iter_rows(values_only=True)
    )

    bitacora._guardar_operacion_excel(tmp_path, _operacion(date(2026, 8, 3), 10))
    filas_despues = list(
        load_workbook(bitacora._ruta_excel(tmp_path))[bitacora._HOJA_OPERACIONES].iter_rows(
            min_row=1, max_row=3, values_only=True
        )
    )
    assert filas_antes == filas_despues


def test_rescribir_hoja_resumen_no_deja_filas_viejas(tmp_path):
    wb = Workbook()
    wb.active.title = bitacora._HOJA_OPERACIONES
    hoja_vieja = wb.create_sheet(bitacora._HOJA_BALANCE_DIARIO, index=1)
    hoja_vieja.append(bitacora._COLUMNAS_BALANCE)
    for i in range(5):
        hoja_vieja.append([f"2026-08-0{i + 1}", 1, 10, 100.0])

    bitacora._rescribir_hoja_resumen(
        wb,
        bitacora._HOJA_BALANCE_DIARIO,
        1,
        bitacora._COLUMNAS_BALANCE,
        [["2026-08-01", 1, 10, 100.0], ["2026-08-02", 1, 20, 100.0]],
    )

    assert wb[bitacora._HOJA_BALANCE_DIARIO].max_row == 3  # encabezado + 2 filas nuevas


def test_orden_de_pestanas_estable_entre_recomputes(tmp_path):
    orden_esperado = [
        bitacora._HOJA_OPERACIONES,
        bitacora._HOJA_BALANCE_DIARIO,
        bitacora._HOJA_BALANCE_SEMANAL,
        bitacora._HOJA_BALANCE_MENSUAL,
        bitacora._HOJA_OPERATIVAS,
    ]
    bitacora._guardar_operacion_excel(tmp_path, _operacion(date(2026, 8, 1), 50))
    assert load_workbook(bitacora._ruta_excel(tmp_path)).sheetnames == orden_esperado

    bitacora._guardar_operacion_excel(tmp_path, _operacion(date(2026, 8, 2), -20))
    bitacora._guardar_operacion_excel(tmp_path, _operacion(date(2026, 8, 3), 15))
    assert load_workbook(bitacora._ruta_excel(tmp_path)).sheetnames == orden_esperado


# --- Resumenes diarios en texto --------------------------------------------

def test_resumen_diario_crea_archivo_en_carpeta_del_mes(tmp_path):
    bitacora._agregar_resumen_nota(tmp_path, date(2026, 8, 2), "hola")
    ruta = tmp_path / "resumenes" / "2026-08" / "2026-08-02.txt"
    assert ruta.exists()
    assert "hola" in ruta.read_text(encoding="utf-8")


def test_resumen_diario_dos_entradas_mismo_dia_no_se_pisan(tmp_path):
    bitacora._agregar_resumen_nota(tmp_path, date(2026, 8, 2), "primera")
    bitacora._agregar_resumen_nota(tmp_path, date(2026, 8, 2), "segunda")
    ruta = tmp_path / "resumenes" / "2026-08" / "2026-08-02.txt"
    lineas = ruta.read_text(encoding="utf-8").splitlines()
    assert len(lineas) == 2
    assert "primera" in lineas[0]
    assert "segunda" in lineas[1]


def test_resumen_diario_cruza_mes_en_carpetas_distintas(tmp_path):
    bitacora._agregar_resumen_nota(tmp_path, date(2026, 7, 31), "fin de julio")
    bitacora._agregar_resumen_nota(tmp_path, date(2026, 8, 1), "inicio agosto")
    assert (tmp_path / "resumenes" / "2026-07" / "2026-07-31.txt").exists()
    assert (tmp_path / "resumenes" / "2026-08" / "2026-08-01.txt").exists()


def test_resumen_operacion_formato_incluye_mercado_y_resultado(tmp_path):
    operacion = _operacion(date(2026, 8, 2), 45.5, mercado="EUR/USD", tipo="Short")
    bitacora._agregar_resumen_operacion(tmp_path, operacion)
    contenido = (tmp_path / "resumenes" / "2026-08" / "2026-08-02.txt").read_text(encoding="utf-8")
    assert "EUR/USD" in contenido
    assert "Short" in contenido
    assert "45.50" in contenido


# --- Entrypoints publicos ---------------------------------------------------

def test_registrar_operacion_feliz(tmp_path, monkeypatch):
    monkeypatch.setattr(bitacora, "_raiz_bitacora", lambda: tmp_path)
    resultado = bitacora.registrar_operacion(
        {"mercado": "BTC/USDT", "tipo": "Long", "resultado_usd": 25, "fecha": "2026-08-02", "operativa": "Breakout"}
    )
    assert "registrada" in resultado.lower()
    assert bitacora._ruta_excel(tmp_path).exists()
    assert (tmp_path / "resumenes" / "2026-08" / "2026-08-02.txt").exists()


def test_registrar_operacion_permission_error_da_mensaje_legible(tmp_path, monkeypatch):
    monkeypatch.setattr(bitacora, "_raiz_bitacora", lambda: tmp_path)

    def _guardar_que_falla(raiz, operacion):
        raise PermissionError("bloqueado")

    monkeypatch.setattr(bitacora, "_guardar_operacion_excel", _guardar_que_falla)

    resultado = bitacora.registrar_operacion({"resultado_usd": 10, "mercado": "BTC", "tipo": "Long"})
    assert "abierto en excel" in resultado.lower()


def test_agregar_nota_bitacora_feliz_no_toca_excel(tmp_path, monkeypatch):
    monkeypatch.setattr(bitacora, "_raiz_bitacora", lambda: tmp_path)
    resultado = bitacora.agregar_nota_bitacora({"nota": "hoy no opere"})
    assert "guardada" in resultado.lower()
    assert not bitacora._ruta_excel(tmp_path).exists()


def test_agregar_nota_bitacora_vacia_no_toca_disco(monkeypatch):
    def _raiz_que_no_deberia_llamarse():
        raise AssertionError("no deberia resolver la raiz con una nota vacia")

    monkeypatch.setattr(bitacora, "_raiz_bitacora", _raiz_que_no_deberia_llamarse)

    resultado = bitacora.agregar_nota_bitacora({"nota": "   "})
    assert "error" in resultado.lower()


# --- Resolver de Escritorio (registro mockeado, nunca el real) -------------

def test_resolver_desktop_usa_registro(monkeypatch):
    monkeypatch.setattr(winreg, "OpenKey", lambda *a, **k: _FakeClave())
    monkeypatch.setattr(winreg, "QueryValueEx", lambda *a, **k: (r"C:\Users\Sebas\Desktop", 1))

    assert bitacora._resolver_desktop() == Path(r"C:\Users\Sebas\Desktop")


def test_resolver_desktop_expande_variables_de_entorno(monkeypatch):
    monkeypatch.setattr(winreg, "OpenKey", lambda *a, **k: _FakeClave())
    monkeypatch.setattr(winreg, "QueryValueEx", lambda *a, **k: (r"%USERPROFILE%\Desktop", 1))
    monkeypatch.setenv("USERPROFILE", r"C:\Users\Sebas")

    assert bitacora._resolver_desktop() == Path(r"C:\Users\Sebas\Desktop")


def test_resolver_desktop_cae_a_home_si_falla_el_registro(monkeypatch):
    def _open_key_que_falla(*a, **k):
        raise OSError("no se pudo abrir la clave")

    monkeypatch.setattr(winreg, "OpenKey", _open_key_que_falla)

    assert bitacora._resolver_desktop() == Path.home() / "Desktop"


def test_raiz_bitacora_compone_sobre_desktop(tmp_path, monkeypatch):
    monkeypatch.setattr(bitacora, "_resolver_desktop", lambda: tmp_path)
    raiz = bitacora._raiz_bitacora()
    assert raiz == tmp_path / "Trade" / "bitacora"
    assert raiz.exists()
