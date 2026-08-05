"""Bitacora de trading de Sebas (operaciones.xlsx + resumenes diarios en
texto).

A diferencia de memoria.py y uso_mensual.py (estado interno de la app, vive
escondido en %APPDATA%), esto NO es estado interno: vive en el Escritorio
real de Windows para que Sebas lo pueda abrir el mismo en el Explorador o en
Excel cuando quiera, sin tener que pedirselo a Trade.
"""
import os
import threading
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

_NOMBRE_ARCHIVO_EXCEL = "operaciones.xlsx"

_HOJA_OPERACIONES = "Operaciones"
_HOJA_BALANCE_DIARIO = "Balance Diario"
_HOJA_BALANCE_SEMANAL = "Balance Semanal"
_HOJA_BALANCE_MENSUAL = "Balance Mensual"
_HOJA_OPERATIVAS = "Operativas"

_COLUMNAS_OPERACIONES = [
    "Fecha", "Hora", "Mercado", "Operativa", "Tipo", "Entrada", "Salida",
    "Tamaño", "Resultado USD", "Resultado R", "Notas",
]
_COLUMNAS_BALANCE = ["Periodo", "Num Operaciones", "P&L Total USD", "Winrate %"]
_COLUMNAS_OPERATIVAS = ["Operativa", "Num Operaciones", "P&L Total USD", "Winrate %"]

# Orden e indice fijo de las hojas de resumen - sin index= explicito, cada
# recompute (borrar + recrear) las va empujando al final y la pestaña de
# Excel termina reordenandose sola de forma visible entre un guardado y el
# siguiente.
_HOJAS_RESUMEN = [
    (_HOJA_BALANCE_DIARIO, 1, _COLUMNAS_BALANCE),
    (_HOJA_BALANCE_SEMANAL, 2, _COLUMNAS_BALANCE),
    (_HOJA_BALANCE_MENSUAL, 3, _COLUMNAS_BALANCE),
    (_HOJA_OPERATIVAS, 4, _COLUMNAS_OPERATIVAS),
]

# Perdidas seguidas en el dia a partir de las cuales registrar_operacion
# agrega una alerta al resultado - ataca directamente la prioridad #2 de
# coaching de Sebas (no tiene un disparador automatico para rachas).
_UMBRAL_ALERTA_RACHA = 2

# Protege el ciclo cargar->modificar->guardar del excel - resguardo barato
# (hoy _llamar_api solo corre en un hilo a la vez, ver main.py), mismo
# patron que _lock_uso en uso_mensual.py.
_lock_bitacora = threading.Lock()


def _resolver_desktop() -> Path:
    """Resuelve el Escritorio real de Windows via el registro - OneDrive
    puede redirigirlo fuera de ~/Desktop, y asumir esa ruta a secas dejaria
    la bitacora en un lugar que Sebas no ve al abrir su escritorio real.
    Import de winreg adentro de la funcion (no al tope del modulo) para que
    un entorno sin registro accesible (o sin winreg, fuera de Windows) no
    rompa el import del modulo entero."""
    try:
        import winreg
        with winreg.OpenKey(
            winreg.HKEY_CURRENT_USER,
            r"Software\Microsoft\Windows\CurrentVersion\Explorer\User Shell Folders",
        ) as clave:
            valor, _ = winreg.QueryValueEx(clave, "Desktop")
        return Path(os.path.expandvars(valor))
    except (ImportError, OSError, FileNotFoundError):
        return Path.home() / "Desktop"


def _raiz_bitacora() -> Path:
    base = _resolver_desktop() / "Trade" / "bitacora"
    base.mkdir(parents=True, exist_ok=True)
    return base


def _ruta_excel(raiz: Path) -> Path:
    return raiz / _NOMBRE_ARCHIVO_EXCEL


def _cargar_o_crear_workbook(ruta: Path) -> Workbook:
    if ruta.exists():
        wb = load_workbook(ruta)
        if _HOJA_OPERACIONES not in wb.sheetnames:
            # Archivo existente pero sin la hoja principal (Sebas pudo
            # borrarla sin querer en Excel) - se recrea vacia en vez de
            # fallar, para no perder el resto del archivo.
            hoja = wb.create_sheet(_HOJA_OPERACIONES, index=0)
            hoja.append(_COLUMNAS_OPERACIONES)
    else:
        wb = Workbook()
        hoja = wb.active
        hoja.title = _HOJA_OPERACIONES
        hoja.append(_COLUMNAS_OPERACIONES)

    for nombre_hoja, indice, columnas in _HOJAS_RESUMEN:
        if nombre_hoja not in wb.sheetnames:
            hoja = wb.create_sheet(nombre_hoja, index=indice)
            hoja.append(columnas)

    return wb


def _agregar_fila_operacion(wb: Workbook, operacion: dict) -> None:
    hoja = wb[_HOJA_OPERACIONES]
    hoja.append([
        operacion["fecha"],
        operacion["hora"],
        operacion["mercado"],
        operacion["operativa"],
        operacion["tipo"],
        operacion["entrada"],
        operacion["salida"],
        operacion["tamano"],
        operacion["resultado_usd"],
        operacion["resultado_r"],
        operacion["notas"],
    ])


def _leer_operaciones(wb: Workbook) -> list:
    """Lee todas las filas de Operaciones ya normalizadas (fecha como
    date, no datetime - openpyxl suele devolver datetime.datetime aunque se
    haya guardado un date, y si no se normaliza aca la agrupacion por dia/
    semana/mes se rompe en silencio)."""
    hoja = wb[_HOJA_OPERACIONES]
    operaciones = []
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if fila[0] is None:
            continue
        fecha = fila[0]
        if isinstance(fecha, datetime):
            fecha = fecha.date()
        operaciones.append({
            "fecha": fecha,
            "hora": fila[1],
            "mercado": fila[2],
            "operativa": fila[3],
            "tipo": fila[4],
            "entrada": fila[5],
            "salida": fila[6],
            "tamano": fila[7],
            "resultado_usd": fila[8],
            "resultado_r": fila[9],
            "notas": fila[10],
        })
    return operaciones


def _racha_perdidas_consecutivas(operaciones: list, fecha: date) -> int:
    """Cuenta la racha de perdidas consecutivas del dia dado (dia
    calendario - "en una sesion" del perfil de Sebas se interpreta asi, sin
    manejo de zona horaria, igual que el resto del modulo), contando hacia
    atras desde la operacion mas reciente de ese dia (ordenadas por hora,
    que es siempre "HH:MM" y por lo tanto ordena bien como texto) hasta la
    primera que no sea perdida - una ganancia, un empate en 0, o un None
    (fila editada a mano por Sebas en Excel) cortan la racha igual."""
    del_dia = [op for op in operaciones if op["fecha"] == fecha]
    del_dia.sort(key=lambda op: op["hora"] or "")

    racha = 0
    for op in reversed(del_dia):
        resultado = op["resultado_usd"]
        if resultado is not None and resultado < 0:
            racha += 1
        else:
            break
    return racha


def _racha_perdidas_del_dia(raiz: Path, fecha: date) -> int:
    """Envoltorio delgado: relee el excel ya guardado en disco y aplica
    _racha_perdidas_consecutivas. Relectura independiente de la que ya hace
    _guardar_operacion_excel al recalcular balances - aceptado por el
    volumen real de Sebas (5-6 operaciones/mes), no vale la pena acoplar el
    contrato de retorno de esa funcion para evitarlo."""
    ruta = _ruta_excel(raiz)
    wb = _cargar_o_crear_workbook(ruta)
    operaciones = _leer_operaciones(wb)
    return _racha_perdidas_consecutivas(operaciones, fecha)


def _clave_dia(operacion: dict) -> str:
    return operacion["fecha"].isoformat()


def _clave_semana(operacion: dict) -> str:
    # Python 3.8: date.isocalendar() devuelve una tupla plana (no tiene
    # .week/.year como en 3.9+) - el proyecto esta fijado a 3.8.
    iso = operacion["fecha"].isocalendar()
    return f"{iso[0]}-W{iso[1]:02d}"


def _clave_mes(operacion: dict) -> str:
    fecha = operacion["fecha"]
    return f"{fecha.year:04d}-{fecha.month:02d}"


def _agrupar_balance(operaciones: list, clave_fn) -> list:
    """Agrupa por la clave dada (dia/semana/mes) y devuelve filas
    [periodo, num_operaciones, pnl_total, winrate] ordenadas por periodo."""
    grupos = {}
    for operacion in operaciones:
        clave = clave_fn(operacion)
        grupo = grupos.setdefault(clave, {"num": 0, "pnl": 0.0, "ganadas": 0})
        resultado = operacion["resultado_usd"] or 0.0
        grupo["num"] += 1
        grupo["pnl"] += resultado
        if resultado > 0:
            grupo["ganadas"] += 1

    filas = []
    for clave in sorted(grupos):
        grupo = grupos[clave]
        winrate = (grupo["ganadas"] / grupo["num"] * 100) if grupo["num"] else 0.0
        filas.append([clave, grupo["num"], round(grupo["pnl"], 2), round(winrate, 1)])
    return filas


def _agrupar_operativas(operaciones: list) -> list:
    """Igual que _agrupar_balance pero agrupando por texto libre de
    operativa (sin normalizar mayusculas/espacios extra mas alla del strip
    ya aplicado al guardar) y ordenado por P&L descendente - para que Sebas
    vea primero sus mejores/peores setups de un vistazo."""
    grupos = {}
    for operacion in operaciones:
        clave = operacion["operativa"] or "Sin especificar"
        grupo = grupos.setdefault(clave, {"num": 0, "pnl": 0.0, "ganadas": 0})
        resultado = operacion["resultado_usd"] or 0.0
        grupo["num"] += 1
        grupo["pnl"] += resultado
        if resultado > 0:
            grupo["ganadas"] += 1

    filas = []
    for clave, grupo in grupos.items():
        winrate = (grupo["ganadas"] / grupo["num"] * 100) if grupo["num"] else 0.0
        filas.append([clave, grupo["num"], round(grupo["pnl"], 2), round(winrate, 1)])
    filas.sort(key=lambda fila: fila[2], reverse=True)
    return filas


def _rescribir_hoja_resumen(wb: Workbook, nombre: str, indice: int, columnas: list, filas: list) -> None:
    if nombre in wb.sheetnames:
        del wb[nombre]
    hoja = wb.create_sheet(nombre, index=indice)
    hoja.append(columnas)
    for fila in filas:
        hoja.append(fila)


def _recalcular_balances(wb: Workbook) -> None:
    operaciones = _leer_operaciones(wb)
    _rescribir_hoja_resumen(
        wb, _HOJA_BALANCE_DIARIO, 1, _COLUMNAS_BALANCE,
        _agrupar_balance(operaciones, _clave_dia),
    )
    _rescribir_hoja_resumen(
        wb, _HOJA_BALANCE_SEMANAL, 2, _COLUMNAS_BALANCE,
        _agrupar_balance(operaciones, _clave_semana),
    )
    _rescribir_hoja_resumen(
        wb, _HOJA_BALANCE_MENSUAL, 3, _COLUMNAS_BALANCE,
        _agrupar_balance(operaciones, _clave_mes),
    )
    _rescribir_hoja_resumen(
        wb, _HOJA_OPERATIVAS, 4, _COLUMNAS_OPERATIVAS,
        _agrupar_operativas(operaciones),
    )


def _guardar_operacion_excel(raiz: Path, operacion: dict) -> None:
    ruta = _ruta_excel(raiz)
    with _lock_bitacora:
        wb = _cargar_o_crear_workbook(ruta)
        _agregar_fila_operacion(wb, operacion)
        _recalcular_balances(wb)
        wb.save(ruta)


def _ruta_resumen_diario(raiz: Path, fecha: date) -> Path:
    carpeta_mes = raiz / "resumenes" / f"{fecha.year:04d}-{fecha.month:02d}"
    carpeta_mes.mkdir(parents=True, exist_ok=True)
    return carpeta_mes / f"{fecha.isoformat()}.txt"


def _agregar_linea_resumen(raiz: Path, fecha: date, linea: str) -> None:
    ruta = _ruta_resumen_diario(raiz, fecha)
    with open(ruta, "a", encoding="utf-8") as archivo:
        archivo.write(linea + "\n")


def _agregar_resumen_operacion(raiz: Path, operacion: dict) -> None:
    signo = "+" if operacion["resultado_usd"] >= 0 else ""
    linea = (
        f"[{operacion['hora']}] {operacion['mercado']} {operacion['tipo']} -> "
        f"{signo}{operacion['resultado_usd']:.2f} USD ({operacion['operativa']})"
    )
    _agregar_linea_resumen(raiz, operacion["fecha"], linea)


def _agregar_resumen_nota(raiz: Path, fecha: date, nota: str) -> None:
    hora = datetime.now().strftime("%H:%M")
    _agregar_linea_resumen(raiz, fecha, f"[{hora}] Nota: {nota}")


def _a_numero_opcional(valor) -> float:
    if valor is None or valor == "":
        return None
    try:
        return float(valor)
    except (TypeError, ValueError):
        return None


def _normalizar_operacion(entrada: dict) -> dict:
    if "resultado_usd" not in entrada or entrada["resultado_usd"] in (None, ""):
        raise ValueError("Falta resultado_usd - no se puede registrar una operacion sin resultado.")
    try:
        resultado_usd = float(entrada["resultado_usd"])
    except (TypeError, ValueError):
        raise ValueError("resultado_usd tiene que ser un numero.")

    fecha_texto = entrada.get("fecha")
    if fecha_texto:
        try:
            fecha = date.fromisoformat(fecha_texto)
        except ValueError:
            raise ValueError(f"Fecha invalida: {fecha_texto!r} (formato esperado YYYY-MM-DD).")
    else:
        fecha = date.today()

    return {
        "fecha": fecha,
        "hora": datetime.now().strftime("%H:%M"),
        "mercado": str(entrada.get("mercado", "")).strip(),
        "operativa": str(entrada.get("operativa", "")).strip() or "Sin especificar",
        "tipo": str(entrada.get("tipo", "")).strip(),
        "entrada": _a_numero_opcional(entrada.get("entrada")),
        "salida": _a_numero_opcional(entrada.get("salida")),
        "tamano": _a_numero_opcional(entrada.get("tamano")),
        "resultado_usd": resultado_usd,
        "resultado_r": _a_numero_opcional(entrada.get("resultado_r")),
        "notas": str(entrada.get("notas", "")).strip(),
    }


def registrar_operacion(entrada: dict) -> str:
    """Manejador de la herramienta registrar_operacion (ver config.py) -
    guarda la fila en el Excel, recalcula los balances, agrega la linea al
    resumen del dia y, si la operacion fue una perdida, chequea la racha de
    perdidas seguidas de hoy y agrega una alerta si llega al umbral. Nunca
    deja pasar una excepcion cruda al loop de tool-use de main.py, siempre
    devuelve un string legible."""
    try:
        operacion = _normalizar_operacion(entrada)
        raiz = _raiz_bitacora()
        _guardar_operacion_excel(raiz, operacion)
        _agregar_resumen_operacion(raiz, operacion)
    except ValueError as error:
        return f"Error: {error}"
    except PermissionError:
        return (
            "Error: no se pudo guardar en operaciones.xlsx - probablemente esta "
            "abierto en Excel. Cierralo e intenta de nuevo."
        )
    except OSError as error:
        return f"Error de archivo: {error}"

    resultado = (
        f"Operacion registrada: {operacion['mercado']} {operacion['tipo']} "
        f"({operacion['resultado_usd']:+.2f} USD) - {operacion['fecha'].isoformat()}."
    )

    if operacion["resultado_usd"] < 0:
        try:
            racha = _racha_perdidas_del_dia(raiz, operacion["fecha"])
        except OSError:
            # La operacion YA se guardo bien - un fallo en esta relectura
            # best-effort no debe pisar la confirmacion de exito de arriba.
            racha = 0
        if racha >= _UMBRAL_ALERTA_RACHA:
            resultado += (
                f" Atencion: van {racha} perdidas seguidas hoy - tu diagnostico identifico "
                "esto como el momento de una pausa mecanica (ej. 15 min tras 2 seguidas, "
                "cortar el dia tras 3) en vez de seguir operando con la cabeza caliente."
            )

    return resultado


def agregar_nota_bitacora(entrada: dict) -> str:
    """Manejador de la herramienta agregar_nota_bitacora (ver config.py) -
    agrega una linea libre al resumen del dia de hoy, sin tocar el Excel."""
    nota = str(entrada.get("nota", "")).strip()
    if not nota:
        return "Error: la nota viene vacia, no se guardo nada."
    try:
        raiz = _raiz_bitacora()
        _agregar_resumen_nota(raiz, date.today(), nota)
        return "Nota guardada en la bitacora de hoy."
    except PermissionError:
        return "Error: no se pudo escribir el resumen del dia (archivo en uso)."
    except OSError as error:
        return f"Error de archivo: {error}"
