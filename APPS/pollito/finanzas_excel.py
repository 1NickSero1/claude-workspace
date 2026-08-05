"""Conecta la skill de Finanzas ("Monito") con el excel real de Sofi
(hojas RESUMEN/GASTOS/TARJETA/DEUDAS - ver plantilla original) para que ella
nunca lo edite a mano: Monito lee su estado financiero, anota gastos nuevos,
actualiza el ingreso del mes y arma el archivo del mes siguiente.

El archivo vive en la PC de Sofi, en la carpeta que ella elija (boton "📊" en
skills/finanzas.py) - solo se recuerda la ruta (%APPDATA%/Pollito), nunca se
embebe ningun dato financiero real en el codigo de la app.

Ningun numero derivado (saldo disponible, cuota mensual de tarjeta, etc.) se
lee del cache de formulas de Excel: si Monito acaba de escribir el archivo,
ese cache ni existe (openpyxl no recalcula formulas). Se recalculan siempre
en Python a partir de los valores crudos.
"""
import json
import os
import shutil
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional

import openpyxl
from openpyxl.cell.cell import MergedCell

HOJAS_ESPERADAS = {"RESUMEN", "GASTOS", "TARJETA", "DEUDAS"}

# Rangos de filas de datos de cada hoja (ver plantilla original) - el resto
# de cada hoja son encabezados, formulas de resumen o el bloque de
# configuracion de TARJETA (H2:I6).
FILA_INICIO_GASTOS, FILA_FIN_GASTOS = 6, 53
FILA_INICIO_DEUDAS, FILA_FIN_DEUDAS = 7, 56
FILA_INICIO_TARJETA, FILA_FIN_TARJETA = 9, 108

EXCEL_FINANZAS_TOOL = {
    "name": "excel_finanzas",
    "description": (
        "Lee y modifica el excel de finanzas personales de Sofi (hojas "
        "RESUMEN/GASTOS/TARJETA/DEUDAS). Ella nunca edita el archivo a mano "
        "- todo pasa por vos. Usa 'leer_estado' antes de responder cualquier "
        "pregunta sobre su plata real (nunca inventes ni asumas numeros). "
        "Usa 'anotar_gasto' cada vez que te cuente un gasto nuevo. Usa "
        "'fijar_ingreso_mes' si te dice cuanto gano este mes. Usa "
        "'crear_archivo_mes' solo cuando te pida explicitamente armar el "
        "archivo de un mes nuevo (arranca GASTOS vacio, pero deuda y "
        "tarjeta pasan igual porque son cuotas de varios meses)."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "accion": {
                "type": "string",
                "enum": ["leer_estado", "anotar_gasto", "fijar_ingreso_mes", "crear_archivo_mes"],
                "description": "Que operacion hacer sobre el excel.",
            },
            "categoria": {
                "type": "string",
                "description": (
                    "Solo para anotar_gasto - ej. Transporte, Alimentacion, "
                    "Salud, Arriendo, Servicios, Otro."
                ),
            },
            "descripcion": {
                "type": "string",
                "description": "Solo para anotar_gasto - detalle corto del gasto (ej. GASOLINA, MERCADO).",
            },
            "monto": {
                "type": "number",
                "description": "Monto en pesos - para anotar_gasto (el gasto) o fijar_ingreso_mes (el ingreso).",
            },
            "estado_pago": {
                "type": "string",
                "enum": ["PAGADO", "PENDIENTE"],
                "description": "Solo para anotar_gasto - default PAGADO si no se aclara.",
            },
            "fecha_pago": {
                "type": "string",
                "description": "Solo para anotar_gasto con estado_pago PENDIENTE - fecha en formato YYYY-MM-DD.",
            },
            "mes_nuevo": {
                "type": "string",
                "description": "Solo para crear_archivo_mes - nombre del mes en espanol, ej. SEPTIEMBRE.",
            },
            "anio_nuevo": {
                "type": "integer",
                "description": "Solo para crear_archivo_mes - ej. 2026.",
            },
            "ingreso_mes_nuevo": {
                "type": "number",
                "description": "Opcional, solo para crear_archivo_mes - si ya te dijo cuanto gana este mes.",
            },
        },
        "required": ["accion"],
    },
}


def _ruta_config() -> Path:
    base = Path(os.environ.get("APPDATA", str(Path.home()))) / "Pollito"
    base.mkdir(parents=True, exist_ok=True)
    return base / "finanzas_excel.json"


def obtener_archivo_actual() -> Optional[Path]:
    ruta_config = _ruta_config()
    if not ruta_config.exists():
        return None
    try:
        datos = json.loads(ruta_config.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None
    archivo = datos.get("archivo_actual")
    if not archivo:
        return None
    ruta = Path(archivo)
    return ruta if ruta.exists() else None


def fijar_archivo_actual(ruta: Path) -> None:
    _ruta_config().write_text(json.dumps({"archivo_actual": str(ruta)}), encoding="utf-8")


def validar_formato(ruta: Path) -> None:
    """Lanza ValueError si el archivo no tiene las 4 hojas esperadas -
    se llama al conectar el archivo por primera vez (ver skills/finanzas.py)
    para no descubrir el problema recien en medio de una conversacion."""
    wb = openpyxl.load_workbook(ruta, read_only=True)
    try:
        faltantes = HOJAS_ESPERADAS - set(wb.sheetnames)
        if faltantes:
            raise ValueError("le faltan las hojas {}".format(", ".join(sorted(faltantes))))
    finally:
        wb.close()


def _limpiar_celda(ws, fila: int, columna: int) -> None:
    """Vacia una celda de forma segura para celdas combinadas: las que no
    son la esquina superior izquierda de un merge son MergedCell (de solo
    lectura en openpyxl), asignarles .value tira AttributeError - se
    saltean sin hacer nada, ya vienen vacias porque solo la esquina del
    merge guarda contenido real."""
    celda = ws.cell(row=fila, column=columna)
    if not isinstance(celda, MergedCell):
        celda.value = None


def _numero(valor) -> float:
    return valor if isinstance(valor, (int, float)) else 0.0


def _texto(valor) -> str:
    return "" if valor is None else str(valor)


def _formatear_fecha(valor) -> str:
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")
    return _texto(valor)


def _leer_gastos(ws) -> list:
    filas = []
    for fila in range(FILA_INICIO_GASTOS, FILA_FIN_GASTOS + 1):
        monto = ws.cell(row=fila, column=4).value
        if not isinstance(monto, (int, float)):
            continue
        filas.append({
            "fila": fila,
            "categoria": _texto(ws.cell(row=fila, column=2).value),
            "descripcion": _texto(ws.cell(row=fila, column=3).value),
            "monto": monto,
            "estado": _texto(ws.cell(row=fila, column=5).value) or "PAGADO",
            "fecha_pago": _formatear_fecha(ws.cell(row=fila, column=6).value),
        })
    return filas


def _leer_deudas(ws) -> list:
    filas = []
    for fila in range(FILA_INICIO_DEUDAS, FILA_FIN_DEUDAS + 1):
        deuda_total = ws.cell(row=fila, column=4).value
        if not isinstance(deuda_total, (int, float)):
            continue
        ya_pagado = _numero(ws.cell(row=fila, column=5).value)
        filas.append({
            "fila": fila,
            "nombre": _texto(ws.cell(row=fila, column=2).value),
            "descripcion": _texto(ws.cell(row=fila, column=3).value),
            "deuda_total": deuda_total,
            "ya_pagado": ya_pagado,
            "saldo_pendiente": deuda_total - ya_pagado,
            "cuota_mes": _numero(ws.cell(row=fila, column=7).value),
            "estado": _texto(ws.cell(row=fila, column=10).value) or "ACTIVA",
        })
    return filas


def _leer_tarjeta(ws) -> list:
    filas = []
    for fila in range(FILA_INICIO_TARJETA, FILA_FIN_TARJETA + 1):
        monto_total = ws.cell(row=fila, column=3).value
        if not isinstance(monto_total, (int, float)):
            continue
        cuotas = ws.cell(row=fila, column=4).value
        cuotas = cuotas if isinstance(cuotas, (int, float)) and cuotas > 0 else 1
        filas.append({
            "fila": fila,
            "fecha": _formatear_fecha(ws.cell(row=fila, column=1).value),
            "descripcion": _texto(ws.cell(row=fila, column=2).value),
            "monto_total": monto_total,
            "cuotas": cuotas,
            "cuota_mensual": monto_total / cuotas,
        })
    return filas


def leer_estado_financiero(ruta: Path) -> str:
    wb = openpyxl.load_workbook(ruta, data_only=False)
    try:
        ws_resumen = wb["RESUMEN"]
        ws_gastos = wb["GASTOS"]
        ws_tarjeta = wb["TARJETA"]
        ws_deudas = wb["DEUDAS"]

        mes_actual = _texto(ws_resumen["A2"].value) or "(sin definir)"
        balance_mes_anterior = _numero(ws_resumen["B13"].value)
        ingreso_mes = _numero(ws_gastos["C2"].value)

        gastos = _leer_gastos(ws_gastos)
        total_gastos = sum(g["monto"] for g in gastos)
        total_pagado = sum(g["monto"] for g in gastos if g["estado"] == "PAGADO")
        total_pendiente = sum(g["monto"] for g in gastos if g["estado"] == "PENDIENTE")

        tarjeta = _leer_tarjeta(ws_tarjeta)
        cuota_mensual_tarjeta = sum(c["cuota_mensual"] for c in tarjeta)

        deudas = _leer_deudas(ws_deudas)
        deudas_activas = [d for d in deudas if d["estado"] == "ACTIVA"]
        cuota_mensual_deudas = sum(d["cuota_mes"] for d in deudas_activas)
        deuda_pendiente_total = sum(d["saldo_pendiente"] for d in deudas_activas)

        saldo_disponible_estimado = (
            balance_mes_anterior + ingreso_mes - total_pagado
            - cuota_mensual_tarjeta - cuota_mensual_deudas
        )

        lineas = [
            "MES: {}".format(mes_actual),
            "Balance mes anterior: {:,.0f}".format(balance_mes_anterior),
            "Ingreso de este mes: {:,.0f}".format(ingreso_mes),
            "",
            "GASTOS ({} anotados) - total {:,.0f}, pagado {:,.0f}, pendiente {:,.0f}:".format(
                len(gastos), total_gastos, total_pagado, total_pendiente
            ),
        ]
        for g in gastos:
            extra = " | vence {}".format(g["fecha_pago"]) if g["fecha_pago"] else ""
            lineas.append(
                "  - {} | {} | {:,.0f} | {}{}".format(
                    g["categoria"], g["descripcion"], g["monto"], g["estado"], extra
                )
            )

        lineas.append("")
        lineas.append(
            "TARJETA DE CREDITO ({} compras activas) - cuota mensual total {:,.0f}:".format(
                len(tarjeta), cuota_mensual_tarjeta
            )
        )
        for c in tarjeta:
            lineas.append(
                "  - {} | {} | total {:,.0f} en {} cuotas ({:,.0f}/mes)".format(
                    c["fecha"], c["descripcion"], c["monto_total"], c["cuotas"], c["cuota_mensual"]
                )
            )

        lineas.append("")
        lineas.append(
            "DEUDAS ({} activas) - saldo pendiente total {:,.0f}, cuota mensual total {:,.0f}:".format(
                len(deudas_activas), deuda_pendiente_total, cuota_mensual_deudas
            )
        )
        for d in deudas:
            lineas.append(
                "  - {} | {} | saldo {:,.0f} de {:,.0f} | cuota {:,.0f}/mes | {}".format(
                    d["nombre"], d["descripcion"], d["saldo_pendiente"], d["deuda_total"],
                    d["cuota_mes"], d["estado"],
                )
            )

        lineas.append("")
        lineas.append("SALDO DISPONIBLE ESTIMADO este mes: {:,.0f}".format(saldo_disponible_estimado))
        lineas.append("(balance anterior + ingreso - gastos pagados - cuota tarjeta - cuota deudas activas)")

        return "\n".join(lineas)
    finally:
        wb.close()


def anotar_gasto(
    ruta: Path,
    categoria: str,
    descripcion: str,
    monto: float,
    estado: str = "PAGADO",
    fecha_pago: Optional[str] = None,
) -> str:
    wb = openpyxl.load_workbook(ruta)
    try:
        ws = wb["GASTOS"]
        fila_libre = next(
            (
                fila for fila in range(FILA_INICIO_GASTOS, FILA_FIN_GASTOS + 1)
                if not isinstance(ws.cell(row=fila, column=4).value, (int, float))
            ),
            None,
        )
        if fila_libre is None:
            return "Error: no hay mas filas libres en GASTOS este mes (limite {}).".format(
                FILA_FIN_GASTOS - FILA_INICIO_GASTOS + 1
            )

        ws.cell(row=fila_libre, column=2, value=categoria)
        ws.cell(row=fila_libre, column=3, value=descripcion)
        ws.cell(row=fila_libre, column=4, value=monto)
        ws.cell(row=fila_libre, column=5, value=estado)
        if fecha_pago:
            ws.cell(row=fila_libre, column=6, value=fecha_pago)
        wb.save(ruta)
        return "Gasto anotado: {} - {} - {:,.0f} ({})".format(categoria, descripcion, monto, estado)
    finally:
        wb.close()


def fijar_ingreso_mes(ruta: Path, monto: float) -> str:
    wb = openpyxl.load_workbook(ruta)
    try:
        wb["GASTOS"]["C2"] = monto
        wb.save(ruta)
        return "Ingreso del mes actualizado a {:,.0f}".format(monto)
    finally:
        wb.close()


def crear_archivo_mes(
    ruta_actual: Path,
    mes_nuevo: str,
    anio_nuevo: int,
    ingreso_mes_nuevo: Optional[float] = None,
) -> str:
    # 1. Calcula el saldo real que queda del mes viejo (mismo calculo que
    #    leer_estado_financiero) para pasarlo como BALANCE MES ANTERIOR.
    wb_vieja = openpyxl.load_workbook(ruta_actual, data_only=False)
    try:
        balance_anterior_previo = _numero(wb_vieja["RESUMEN"]["B13"].value)
        ws_gastos_vieja = wb_vieja["GASTOS"]
        ingreso_mes_viejo = _numero(ws_gastos_vieja["C2"].value)
        total_pagado_viejo = sum(
            g["monto"] for g in _leer_gastos(ws_gastos_vieja) if g["estado"] == "PAGADO"
        )
        cuota_mensual_tarjeta_vieja = sum(c["cuota_mensual"] for c in _leer_tarjeta(wb_vieja["TARJETA"]))
        cuota_mensual_deudas_vieja = sum(
            d["cuota_mes"] for d in _leer_deudas(wb_vieja["DEUDAS"]) if d["estado"] == "ACTIVA"
        )
        saldo_nuevo_balance = (
            balance_anterior_previo + ingreso_mes_viejo - total_pagado_viejo
            - cuota_mensual_tarjeta_vieja - cuota_mensual_deudas_vieja
        )
    finally:
        wb_vieja.close()

    # 2. El archivo nuevo arranca como una COPIA exacta del actual (asi
    #    TARJETA/DEUDAS pasan tal cual, formulas y formato incluido, ya que
    #    son cuotas activas de varios meses, no datos de un mes solo).
    nombre_nuevo = "ECONOMIA_MENSUAL_{}_{}.xlsx".format(mes_nuevo.upper(), anio_nuevo)
    ruta_nueva = ruta_actual.parent / nombre_nuevo
    if ruta_nueva.exists():
        return "Error: ya existe un archivo para {} de {} ({}).".format(mes_nuevo, anio_nuevo, nombre_nuevo)

    shutil.copy2(ruta_actual, ruta_nueva)

    wb_nueva = openpyxl.load_workbook(ruta_nueva)
    try:
        ws_resumen = wb_nueva["RESUMEN"]
        ws_resumen["A2"] = "{} DE {}".format(mes_nuevo.upper(), anio_nuevo)
        ws_resumen["B13"] = round(saldo_nuevo_balance)

        ws_gastos = wb_nueva["GASTOS"]
        ws_gastos["C2"] = ingreso_mes_nuevo if ingreso_mes_nuevo is not None else 0
        for fila in range(FILA_INICIO_GASTOS, FILA_FIN_GASTOS + 1):
            for columna in (2, 3, 4, 5, 6):  # categoria/descripcion/monto/estado/fecha
                _limpiar_celda(ws_gastos, fila, columna)
            # Pareja el numerado (columna A) al mismo esquema de formula que
            # ya usaban las filas 16+ - las primeras 10 lo tenian quemado a
            # mano de una carga anterior, queda consistente de ahi en mas.
            ws_gastos.cell(row=fila, column=1).value = '=IF(D{r}<>"",ROW()-5,"")'.format(r=fila)
        # Bloc de cuentas sueltas del mes viejo (SUELDO/DEBE HABER/etc, a la
        # derecha de la tabla de gastos) - no es parte de la plantilla real,
        # son cuentas mentales de ese mes especifico. Tiene celdas combinadas
        # (G6:H6, I9:J9, etc) - _limpiar_celda las saltea.
        for fila in range(6, 20):
            for columna in range(7, 13):  # G..L
                _limpiar_celda(ws_gastos, fila, columna)

        wb_nueva.save(ruta_nueva)
    finally:
        wb_nueva.close()

    fijar_archivo_actual(ruta_nueva)

    mensaje = (
        "Archivo creado para {} de {}: {}. Pase el balance que quedaba del "
        "mes anterior ({:,.0f}) y deje TARJETA/DEUDAS igual - GASTOS arranca "
        "vacio.".format(mes_nuevo, anio_nuevo, nombre_nuevo, saldo_nuevo_balance)
    )
    if ingreso_mes_nuevo is None:
        mensaje += " Todavia falta el ingreso de este mes - pedile el monto y usa fijar_ingreso_mes."
    return mensaje


def crear_manejador_excel() -> Callable[[dict], str]:
    def ejecutar(entrada: dict) -> str:
        accion = entrada.get("accion")
        ruta = obtener_archivo_actual()
        if ruta is None:
            return (
                'Todavia no hay ningun excel conectado. Pedile a Sofi que '
                'toque el boton "📊" (arriba del cuadro de mensaje) para '
                'elegir su archivo.'
            )

        try:
            if accion == "leer_estado":
                return leer_estado_financiero(ruta)

            if accion == "anotar_gasto":
                monto = entrada.get("monto")
                if not isinstance(monto, (int, float)):
                    return "Error: falta el monto del gasto."
                return anotar_gasto(
                    ruta,
                    categoria=entrada.get("categoria") or "Otro",
                    descripcion=entrada.get("descripcion") or "",
                    monto=monto,
                    estado=entrada.get("estado_pago") or "PAGADO",
                    fecha_pago=entrada.get("fecha_pago"),
                )

            if accion == "fijar_ingreso_mes":
                monto = entrada.get("monto")
                if not isinstance(monto, (int, float)):
                    return "Error: falta el monto del ingreso."
                return fijar_ingreso_mes(ruta, monto)

            if accion == "crear_archivo_mes":
                mes_nuevo = entrada.get("mes_nuevo")
                anio_nuevo = entrada.get("anio_nuevo")
                if not mes_nuevo or not isinstance(anio_nuevo, int):
                    return "Error: falta el mes y/o el anio del archivo nuevo."
                return crear_archivo_mes(
                    ruta,
                    mes_nuevo=mes_nuevo,
                    anio_nuevo=anio_nuevo,
                    ingreso_mes_nuevo=entrada.get("ingreso_mes_nuevo"),
                )

            return "Accion de excel desconocida: {}".format(accion)
        except KeyError as error:
            return "Error: el archivo no tiene la hoja esperada {}.".format(error)
        except PermissionError:
            return (
                "Error: no pude guardar el archivo - seguramente esta "
                "abierto en Excel. Pedile a Sofi que lo cierre e intenta "
                "de nuevo."
            )
        except OSError as error:
            return "Error de archivo: {}".format(error)

    return ejecutar
